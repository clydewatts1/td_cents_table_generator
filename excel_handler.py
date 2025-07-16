import pandas as pd
import glob
import os
import re
import openpyxl
import yaml
import logging
import pprint

def validate_ddl_excel_files(config:dict):
    excel_path = config.get('excel_path', '')
    excel_pattern = config.get('excel_pattern', '')
    if not excel_path:
        logging.error("Excel path is not specified in the configuration.",excel_path)
        return
    # glob to find all Excel files in the directory
    excel_files = glob.glob(os.path.join(excel_path, excel_pattern))
    if not excel_files:
        logging.error("No Excel files found in the specified directory.")
        return
    ignore_files = config.get('ignore_files', [])
    if not isinstance(ignore_files, list):
        logging.error("Ignore files is not a list in the configuration.")
        return
    bad_files = []
    for excel_file in excel_files:
        logging.info("Reading Excel file: %s", excel_file)
        # get the file name without path
        file_name = os.path.basename(excel_file)
        # Check if the file is in the ignore list
        if file_name in ignore_files:
            logging.info("Ignoring file: %s", file_name)
            continue
        # Check if the file exists
        # Read the Excel file and convert it to YAML files
        read_excel_file(excel_file, config)
        # get list of sheets in the Excel file
        workbook = openpyxl.load_workbook(excel_file)
        sheets = workbook.sheetnames
        logging.info("Available sheets in the Excel file: %s", sheets)
        # Check if the required sheets are present
        excel_sheets = config.get('excel_sheets', {})
        if not excel_sheets:
            logging.error("No Excel sheets specified in the configuration.")
            bad_files.append(excel_file)
            return
        required_sheets = excel_sheets.get('required_sheets', [])
        if not required_sheets:
            logging.error("No required sheets specified in the configuration.")
            bad_files.append(excel_file)
            return
        missing_sheets = [sheet for sheet in required_sheets if sheet not in sheets]
        if missing_sheets:
            logging.error("The following required sheets are missing in the Excel file: %s", missing_sheets)
            bad_files.append(excel_file+": " + str(missing_sheets))
            continue
        logging.info("All required sheets are present in the Excel file.")
    # print out the bad files
    if bad_files:
        logging.error("The following Excel files are missing required sheets: %s", bad_files)
        fhb = open('badfiles.txt', 'w')
        for file in bad_files:
            # write file badfiles.txt
           fhb.write(f"{file}\n")
        fhb.close()
        logging.info("Bad files written to badfiles.txt")
    else:
        logging.info("All Excel files are valid and contain the required sheets.")




def read_excel_file(file_path,config:dict):
    # sometimes the config comes at the config level , adjust
    if 'config' in config:
        config = config['config']
    # exel_list to config
    if not isinstance(config, dict):
        logging.error("Configuration is not in the expected format. It should be a dictionary.")
        return 
    if 'process_sheets' not in config:
        config['process_sheets'] = {}
    # get file name from file path
    excel_file_name = os.path.basename(file_path)
    config['process_sheets'][excel_file_name] = {}
    config['process_sheets'][excel_file_name]['status'] = 'processing'
    config['process_sheets'][excel_file_name]['file_path'] = file_path
    config['process_sheets'][excel_file_name]['yaml_path'] = config.get('yaml_path', 'yaml_files')
    yaml_filename= f"{os.path.splitext(excel_file_name)[0]}.yaml"
    config['process_sheets'][excel_file_name]['yaml_file'] = yaml_filename

    # Initialize variables
    yaml_file = None
    table_config = {}
    # Placeholder function to read Excel file
    # Print out list of sheets in the Excel file
    logging.info("Reading Excel file: %s", file_path)
    logging.info("Available sheets: %s", openpyxl.load_workbook(file_path).sheetnames)
    # check if required sheets are present in the Excel file
    excel_sheets = config.get('excel_sheets', {})
    if not excel_sheets:
        logging.error("No Excel sheets specified in the configuration.")
        return 
    required_sheets = excel_sheets.get('required_sheets', [])
    if not required_sheets:
        logging.error("No required sheets specified in the configuration.")
        return
    # Load the workbook and check for required sheets
    workbook = openpyxl.load_workbook(file_path)
    sheets = workbook.sheetnames
    logging.info("Available sheets in the Excel file: %s", sheets)
    missing_sheets = [sheet for sheet in required_sheets if sheet not in sheets]
    if missing_sheets:
        logging.error("The following required sheets are missing in the Excel file: %s", missing_sheets)
        return
    # Read BUILD tab in the Excel file to df_build
    df_build = pd.read_excel(file_path, sheet_name='BUILD')
    df_build = df_build.fillna('')
    df_build = df_build.astype(str)
    # Clean up the data by stripping whitespace and removing extra spaces
    df_build = df_build.applymap(lambda x: x.strip())
    df_build = df_build.applymap(lambda x: re.sub(r'\s+', ' ', x))
    # remove all spaces from the column names
    df_build.columns = df_build.columns.str.replace(' ', '')
    # and make lower case
    df_build.columns = df_build.columns.str.lower()

    # remove empty rows
    df_build = df_build[df_build.apply(lambda row: row.astype(str).str.strip().any(), axis=1)]
    # convert to dictionary
    dict_build = df_build.to_dict(orient='records')
    # Print out head 10
    logging.info("Read BUILD sheet: %s", df_build.info()  )
    # Read the Excel file and convert it to a dictionary
    # You can use pandas or openpyxl to read the Excel file and convert it to a dictionary
    df_table = pd.read_excel(file_path,sheet_name='TABLE')
    df_table = df_table.fillna('')
    df_table = df_table.astype(str)
    df_table = df_table.applymap(lambda x: x.strip())
    df_table = df_table.applymap(lambda x: re.sub(r'\s+', ' ', x))

    # Print out head 10
    logging.info("Read TABLE sheet: %s", df_table.head(10))

    # print config
    logging.info("Config: %s", pprint.pformat(config))
    excel_table_config = config['excel_sheets']['TABLE']

    # Get the cell locations for the table configuration
    # Assuming the configuration contains the cell locations for the table configuration
    modelname_cell = excel_table_config.get('modelname_cell', '')
    databasename_cell = excel_table_config.get('databasename_cell', '')
    table_name_cell = excel_table_config.get('table_name_cell', '')
    table_description_cell = excel_table_config.get('table_description_cell', '')
    table_kind_cell = excel_table_config.get('table_kind_cell', '')
    primary_index_cell = excel_table_config.get('primary_index_cell', '')
    table_option_cell = excel_table_config.get('table_option_cell', '')
    control_columns_cell = excel_table_config.get('control_columns_cell', '')
    pattern_cell = excel_table_config.get('patterns_column_cell', '')
    table_quote_cell = excel_table_config.get('table_quote_cell', '')
    #print(databasename_cell,df_table.iat[databasename_cell[0], databasename_cell[1]])
    config_table = {}
    config_table['modelname'] = df_table.iat[modelname_cell[0], modelname_cell[1]]
    config_table['databasename'] = df_table.iat[databasename_cell[0], databasename_cell[1]]
    config_table['table_name'] = df_table.iat[table_name_cell[0], table_name_cell[1]]
    config_table['table_description'] = df_table.iat[table_description_cell[0], table_description_cell[1]]
    config_table['table_description'] = config_table['table_description'].replace("'","''")
    config_table['table_kind'] = df_table.iat[table_kind_cell[0], table_kind_cell[1]]
    config_table['primary_index'] = df_table.iat[primary_index_cell[0], primary_index_cell[1]]
    config_table['table_option'] = df_table.iat[table_option_cell[0], table_option_cell[1]]
    config_table['control_columns'] = df_table.iat[control_columns_cell[0], control_columns_cell[1]]
    config_table['patterns'] = df_table.iat[pattern_cell[0], pattern_cell[1]]
    config_table['table_quote'] = df_table.iat[table_quote_cell[0], table_quote_cell[1]]
    # convert to list
    config_table['patterns'] = config_table['patterns'].split(',')
    #
    column_row = config['excel_sheets']['TABLE']['column_row']
    column_name = config['excel_sheets']['TABLE']['column_name']   
    column_type = config['excel_sheets']['TABLE']['column_type']
    column_format = config['excel_sheets']['TABLE']['column_format']
    column_is_null = config['excel_sheets']['TABLE']['column_is_null']
    column_option = config['excel_sheets']['TABLE']['column_option']
    column_description = config['excel_sheets']['TABLE']['column_description']
    column_compression = config['excel_sheets']['TABLE']['column_compression']
    column_quote = config['excel_sheets']['TABLE']['column_quote']
    # get number of rows in datafram
    num_rows = df_table.shape[0]
    config_table['columns'] = []
    for i in range(column_row,num_rows):
        column_config = {}
        column_config['column_name'] = df_table.iat[i, column_name]
        column_name_value = df_table.iat[i, column_name]
        column_type_value = df_table.iat[i, column_type]
        column_format_value = df_table.iat[i, column_format]
        column_is_null_value = df_table.iat[i, column_is_null]
        column_option_value = df_table.iat[i, column_option]
        column_description_value = df_table.iat[i, column_description]
        column_description_value = column_description_value.replace("'","''")
        column_compression_value = df_table.iat[i, column_compression]
        column_quote_value = df_table.iat[i, column_quote]

        if not isinstance(column_name_value, str) or not isinstance(column_type_value, str):
            logging.error("Column name or type is not a string.")
            continue
        if not isinstance(column_format_value, str) or not isinstance(column_is_null_value, str):
            logging.error("Column format or is_null is not a string.")
            continue
        if not isinstance(column_option_value, str) or not isinstance(column_description_value, str):
            logging.error("Column option or description is not a string.")
            continue
        column_config['column_name'] = column_name_value
        column_config['column_type'] = column_type_value
        # if format value does not start with a single quote, add a single quote
        if not column_format_value.startswith("'"):
            column_format_value = "'" + column_format_value
        # if formart value does not end with a single quote, add a single quote
        if not column_format_value.endswith("'"):
            column_format_value = column_format_value + "'"
        # if format value starts with FORMAT, remove it
        if column_format_value.startswith("FORMAT"):
            column_format_value = column_format_value.replace("FORMAT", "")
        if column_format_value == "''" or column_format_value == "'":
            column_format_value = ''
        column_config['column_format'] = column_format_value
        column_config['is_null'] = column_is_null_value
        column_config['option'] = column_option_value
        column_config['description'] = column_description_value
        column_config['compression'] = column_compression_value
        column_config['quote'] = column_quote_value

        config_table['columns'].append(column_config)
    print("Config table: %s", config_table)
    config_table['BUILD'] = dict_build
    # Save the config to a YAML file
    # Create the directory if it doesn't exist
    if not os.path.exists(config['yaml_path']):
        os.makedirs(config['yaml_path'])
    # Save the config to a YAML file
    yaml_file = os.path.join(config['yaml_path'],yaml_filename)
    with open(yaml_file, 'w') as file:
        yaml.dump(config_table, file, default_flow_style=False, sort_keys=False)
    logging.info("Config saved to: %s", yaml_file)
    return yaml_file

    

def read_excel_files_and_convert_to_yaml_files(config):
    # Assuming the config contains a list of Excel files to read
    excel_path = config.get('excel_path', '')
    excel_pattern = config.get('excel_pattern', '')
    ignore_files = config.get('ignore_files', [])
    if not isinstance(ignore_files, list):
        logging.error("Ignore files is not a list in the configuration.")
        return
    if not excel_path:
        logging.error("Excel path is not specified in the configuration.",excel_path)
        return
    # glob to find all Excel files in the directory
    excel_files = glob.glob(os.path.join(excel_path, excel_pattern))
    if not excel_files:
        logging.error("No Excel files found in the specified directory.")
        return
    for excel_file in excel_files:
        logging.info("Reading Excel file: %s", excel_file)
        # get the file name without path
        file_name = os.path.basename(excel_file)
        # Check if the file is in the ignore list
        if file_name in ignore_files:
            logging.info("Ignoring file: %s", file_name)
            continue
        # Read the Excel file and convert it to YAML files
        read_excel_file(excel_file, config)

